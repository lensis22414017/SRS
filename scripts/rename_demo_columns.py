#!/usr/bin/env python3
"""rename_demo_columns.py — 批量规范化 30 个 demo_sites xlsx 列名

将英文括号 + 斜杠单位 → 中文括号 + 简洁单位（裴总要求）
"""

import os
import shutil
import openpyxl
from pathlib import Path

DEMO_DIR = Path(__file__).resolve().parent.parent / "data" / "demo_sites"

# 列名映射表: 旧列名 → 新列名
COLUMN_RENAME = {
    "pH": "pH",
    "有机质(%)": "有机质（%）",
    "CEC(cmol/kg)": "CEC（cmol/kg）",
    "砂粒(%)": "砂粒（%）",
    "粉粒(%)": "粉粒（%）",
    "黏粒(%)": "黏粒（%）",
    "容重(g/cm³)": "容重（g/cm³）",
    "容重(g/cm3)": "容重（g/cm³）",
    "全氮(g/kg)": "全氮（g/kg）",
    "海拔(m)": "海拔（m）",
    "电导率(mS/cm)": "电导率（mS/cm）",
    "全磷(g/kg)": "全磷（g/kg）",
    "全钾(g/kg)": "全钾（g/kg）",
    "碱解氮(mg/kg)": "碱解氮（mgkg）",
    "速效磷(mg/kg)": "速效磷（mgkg）",
    "速效钾(mg/kg)": "速效钾（mgkg）",
    "镉_Cd(mg/kg)": "镉_Cd（mgkg）",
    "铅_Pb(mg/kg)": "铅_Pb（mgkg）",
    "砷_As(mg/kg)": "砷_As（mgkg）",
    "铬_Cr(mg/kg)": "铬_Cr（mgkg）",
    "汞_Hg(mg/kg)": "汞_Hg（mgkg）",
    "铜_Cu(mg/kg)": "铜_Cu（mgkg）",
    "锌_Zn(mg/kg)": "锌_Zn（mgkg）",
    "镍_Ni(mg/kg)": "镍_Ni（mgkg）",
    "多环芳烃_PAHs(mg/kg)": "多环芳烃_PAHs（mgkg）",
    "苯并芘_BaP(mg/kg)": "苯并芘_BaP（mgkg）",
    "有机氯_OCPs(mg/kg)": "有机氯_OCPs（mgkg）",
    "滴滴涕_DDTs(mg/kg)": "滴滴涕_DDTs（mgkg）",
    "多氯联苯_PCBs(mg/kg)": "多氯联苯_PCBs（mgkg）",
    "六六六_HCHs(mg/kg)": "六六六_HCHs（mgkg）",
    "邻苯二甲酸酯_PAEs(mg/kg)": "邻苯二甲酸酯_PAEs（mgkg）",
    "多溴联苯醚_PBDEs(mg/kg)": "多溴联苯醚_PBDEs（mgkg）",
    "全氟化合物_PFASs(mg/kg)": "全氟化合物_PFASs（mgkg）",
    "总石油烃_TPH(mg/kg)": "总石油烃_TPH（mgkg）",
    "高分子量PAHs(mg/kg)": "高分子量PAHs（mgkg）",
    "低分子量PAHs(mg/kg)": "低分子量PAHs（mgkg）",
    # 元数据列也规范化
    "采样深度(cm)": "采样深度（cm）",
    "年均降水(mm)": "年均降水（mm）",
}


def rename_excel_columns(filepath: Path) -> bool:
    """重写 Excel 表头：只改第一行（表头），数据行不动。"""
    try:
        # 先读为只读获取数据
        wb_read = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws_read = wb_read.active
        rows = list(ws_read.iter_rows(values_only=True))
        wb_read.close()

        if not rows:
            print(f"  [SKIP] {filepath.name}: 空文件")
            return False

        old_header = list(rows[0])
        new_header = []
        renamed_count = 0

        for col_name in old_header:
            col_str = str(col_name) if col_name is not None else ""
            new_name = COLUMN_RENAME.get(col_str, col_str)
            new_header.append(new_name)
            if new_name != col_str:
                renamed_count += 1

        if renamed_count == 0:
            print(f"  [SKIP] {filepath.name}: 列名已是目标格式")
            return False

        # 写回：用备份+覆盖策略
        wb = openpyxl.Workbook()
        ws = wb.active

        # 写表头
        for c, val in enumerate(new_header, 1):
            ws.cell(row=1, column=c, value=val)

        # 写数据行
        for r, row_data in enumerate(rows[1:], 2):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)

        # 保留第一个 sheet 名
        ws.title = "Sheet1"

        # 先写临时文件再替换（避免锁文件问题）
        tmp = filepath.with_suffix(".tmp")
        wb.save(tmp)
        wb.close()
        shutil.move(str(tmp), str(filepath))

        print(f"  [OK] {filepath.name}: 重命名 {renamed_count} 列")
        return True

    except Exception as e:
        print(f"  [FAIL] {filepath.name}: {e}")
        return False


def main():
    xlsx_files = sorted(DEMO_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"目录 {DEMO_DIR} 中无 .xlsx 文件")
        return

    print(f"找到 {len(xlsx_files)} 个 xlsx 文件\n")
    ok, skip, fail = 0, 0, 0

    for fp in xlsx_files:
        if fp.name.startswith("~$"):
            continue
        result = rename_excel_columns(fp)
        if result is True:
            ok += 1
        elif result is False:
            skip += 1

    print(f"\n完成: {ok} 已重命名, {skip} 已跳过")


if __name__ == "__main__":
    main()

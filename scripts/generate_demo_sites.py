# -*- coding: utf-8 -*-
"""
生成 18 份模拟污染场地数据 Excel 文件（用于甲方领导演示）。

输出目录：data/demo_sites/
文件名格式：演示_XX省_类型_严重度.xlsx
  - 类型：HM（重金属）/ OP（有机）/ HMOP（复合）
  - 严重度：轻度 / 中度 / 重度

覆盖矩阵（共 18 份）：
  - 污染类型：HM 6 份、OP 6 份、HMOP 6 份
  - 严重度：轻度/中度/重度 各 6 份（每种类型 × 每种严重度 = 2 份）
  - 省份：覆盖云南、广东、湖南、江苏、浙江、四川、山东、河南、湖北、陕西、辽宁、福建

每份文件列结构与真实数据保持一致：
  序号、采样点编号、经度、纬度、区域、深度_上限(cm)、深度_下限(cm)、
  土壤类型、pH、有机质(g/kg)、全氮(g/kg)、全磷(g/kg)、全钾(g/kg)、
  碱解氮(mg/kg)、速效磷(mg/kg)、速效钾(mg/kg)、
  重金属因子：镉_Cd、铅_Pb、砷_As、铜_Cu、锌_Zn、铬_Cr、汞_Hg、镍_Ni、
  有机因子：苯并芘_BaP、六六六_HCHs、滴滴涕_DDTs、石油烃_TPH、多氯联苯_PCI、
  备注

数值依据：
  - GB15618-2018 农用地土壤污染风险筛选值（mg/kg）：
      Cd=0.6, Pb=170, As=40, Cu=100, Zn=300, Cr=150, Hg=1.3, Ni=100（pH 5.5-8.5）
  - GB36600-2018 建设用地第一类用地筛选值（mg/kg）：
      BaP=0.55, HCHs(α+γ)=6.5（近似），DDTs(总)=1.0（参考），石油烃(C10-C40)=826（一类）, 多氯联苯=0.2
    （为便于演示，OP 场地参照"农用地/居住"较严值：BaP=0.55, HCHs=6.5, DDTs=1.0, 石油烃=100, PCB=0.2）
  - 背景值（mg/kg）：Cu≈20-40, Pb≈20-35, Zn≈60-100, Cr≈60-80, Ni≈25-40,
                    Cd≈0.1-0.3, As≈8-15, Hg≈0.05-0.1
  - 土壤养分正常范围：有机质 1-4% (10-40 g/kg)、全氮 0.1-0.3% (1-3 g/kg)、
                    全磷 0.5-1.5 g/kg、全钾 15-25 g/kg、碱解氮 30-150 mg/kg、
                    速效磷 5-40 mg/kg、速效钾 50-250 mg/kg
  - pH：大部分 6.0-8.0，少量 4.5-5.5

用法：
    python scripts/generate_demo_sites.py
"""

from __future__ import annotations

import os
import random
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# 固定随机种子，保证每次运行结果一致（演示场景可复现）
# -----------------------------------------------------------------------------
SEED = 20250731
random.seed(SEED)

# -----------------------------------------------------------------------------
# 路径
# -----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "demo_sites")

# -----------------------------------------------------------------------------
# 列结构（与真实数据保持一致的命名风格）
# 单位说明写在表头括号内
# -----------------------------------------------------------------------------
COLUMNS: List[str] = [
    "序号",
    "采样点编号",
    "经度",
    "纬度",
    "区域",
    "深度_上限(cm)",
    "深度_下限(cm)",
    "土壤类型",
    "pH",
    "有机质(g/kg)",
    "全氮(g/kg)",
    "全磷(g/kg)",
    "全钾(g/kg)",
    "碱解氮(mg/kg)",
    "速效磷(mg/kg)",
    "速效钾(mg/kg)",
    "镉_Cd(mg/kg)",
    "铅_Pb(mg/kg)",
    "砷_As(mg/kg)",
    "铜_Cu(mg/kg)",
    "锌_Zn(mg/kg)",
    "铬_Cr(mg/kg)",
    "汞_Hg(mg/kg)",
    "镍_Ni(mg/kg)",
    "苯并芘_BaP(mg/kg)",
    "六六六_HCHs(mg/kg)",
    "滴滴涕_DDTs(mg/kg)",
    "石油烃_TPH(mg/kg)",
    "多氯联苯_PCI(mg/kg)",
    "备注",
]

# -----------------------------------------------------------------------------
# 国标筛选值（mg/kg）— 超标判断依据
# GB15618 农用地筛选值（pH 5.5-8.5 区间）
# -----------------------------------------------------------------------------
HM_THRESHOLD: Dict[str, float] = {
    "镉_Cd(mg/kg)": 0.6,
    "铅_Pb(mg/kg)": 170.0,
    "砷_As(mg/kg)": 40.0,
    "铜_Cu(mg/kg)": 100.0,
    "锌_Zn(mg/kg)": 300.0,
    "铬_Cr(mg/kg)": 150.0,
    "汞_Hg(mg/kg)": 1.3,
    "镍_Ni(mg/kg)": 100.0,
}

# GB36600 演示用筛选值（取较严的第一类/农用地相关值）
OP_THRESHOLD: Dict[str, float] = {
    "苯并芘_BaP(mg/kg)": 0.55,
    "六六六_HCHs(mg/kg)": 6.5,
    "滴滴涕_DDTs(mg/kg)": 1.0,
    "石油烃_TPH(mg/kg)": 100.0,
    "多氯联苯_PCI(mg/kg)": 0.2,
}

# 重金属背景值范围（均值, 标准差）—— 未超标时采样于此
HM_BACKGROUND: Dict[str, Tuple[float, float]] = {
    "镉_Cd(mg/kg)": (0.18, 0.08),     # 0.1-0.3
    "铅_Pb(mg/kg)": (25.0, 7.0),      # 20-35
    "砷_As(mg/kg)": (11.0, 3.0),      # 8-15
    "铜_Cu(mg/kg)": (28.0, 7.0),      # 20-40
    "锌_Zn(mg/kg)": (78.0, 18.0),     # 60-100
    "铬_Cr(mg/kg)": (68.0, 10.0),     # 60-80
    "汞_Hg(mg/kg)": (0.07, 0.02),     # 0.05-0.1
    "镍_Ni(mg/kg)": (31.0, 6.0),      # 25-40
}

# 有机物背景值范围（未超标时）—— 多为未检出/痕量
OP_BACKGROUND: Dict[str, Tuple[float, float]] = {
    "苯并芘_BaP(mg/kg)": (0.03, 0.015),    # 接近背景，<0.05
    "六六六_HCHs(mg/kg)": (0.08, 0.04),    # 痕量残留
    "滴滴涕_DDTs(mg/kg)": (0.05, 0.03),
    "石油烃_TPH(mg/kg)": (15.0, 8.0),      # 5-30 自然本底
    "多氯联苯_PCI(mg/kg)": (0.02, 0.01),
}

# 严重度配置：超标比例 & 超标倍数范围（lower, upper）
# 倍数含义：实测值 = 筛选值 × 倍数（再叠加 ±15% 点位间抖动）
# 为保证叠加抖动后仍落在规范区间内（1-3 / 3-10 / 10-50），区间略收窄
SEVERITY_CFG: Dict[str, Dict] = {
    "轻度": {"ratio": 0.30, "mult": (1.0, 2.5)},    # 叠加 ±15% → 约 1-3 倍
    "中度": {"ratio": 0.60, "mult": (3.0, 8.5)},    # 叠加 ±15% → 约 3-10 倍
    "重度": {"ratio": 0.85, "mult": (10.0, 43.0)},  # 叠加 ±15% → 约 10-50 倍
}

# -----------------------------------------------------------------------------
# 18 份文件的覆盖矩阵
# (province, province_code, lon, lat, poll_type, severity)
# poll_type: HM / OP / HMOP
# 省份中心经纬度（取省内典型污染关注城市的代表坐标）
# -----------------------------------------------------------------------------
SITE_MATRIX: List[Dict] = [
    # --- 重金属 HM × 6（云南、湖南、广东、广西... 但要求覆盖指定省份，故选 6 省）---
    {"province": "云南",   "city": "个旧", "code": "YN", "lon": 103.15, "lat": 23.35, "type": "HM",   "severity": "重度"},  # 锡都、Cd/Pb/As
    {"province": "湖南",   "city": "衡阳", "code": "HN", "lon": 112.61, "lat": 26.89, "type": "HM",   "severity": "轻度"},  # 有色金属，Cd/Pb
    {"province": "广东",   "city": "韶关", "code": "GD", "lon": 113.60, "lat": 24.81, "type": "HM",   "severity": "中度"},  # 凡口铅锌矿，Pb/Zn/Cd
    {"province": "四川",   "city": "凉山", "code": "SC", "lon": 102.27, "lat": 27.90, "type": "HM",   "severity": "重度"},  # 攀西有色矿带，Cd/Pb/As
    {"province": "江苏",   "city": "南京", "code": "JS", "lon": 118.78, "lat": 32.06, "type": "HM",   "severity": "轻度"},  # 长江下游
    {"province": "浙江",   "city": "富阳", "code": "ZJ", "lon": 119.96, "lat": 30.05, "type": "HM",   "severity": "中度"},  # Cu/Zn
    # --- 有机 OP × 6 ---
    {"province": "江苏",   "city": "南京栖霞", "code": "JS", "lon": 118.91, "lat": 32.11, "type": "OP",  "severity": "中度"},  # 化工厂原址，BaP/TPH
    {"province": "辽宁",   "city": "沈阳",   "code": "LN", "lon": 123.43, "lat": 41.80, "type": "OP",  "severity": "重度"},  # 石化，TPH/BaP
    {"province": "山东",   "city": "淄博",   "code": "SD", "lon": 118.05, "lat": 36.78, "type": "OP",  "severity": "重度"},  # 石化区，TPH/PCB
    {"province": "湖北",   "city": "武汉",   "code": "HB", "lon": 114.31, "lat": 30.59, "type": "OP",  "severity": "中度"},  # 农药厂原址，HCHs/DDTs
    {"province": "四川",   "city": "成都",   "code": "SC", "lon": 104.07, "lat": 30.67, "type": "OP",  "severity": "轻度"},  # 焦化厂原址，BaP
    {"province": "福建",   "city": "福州",   "code": "FJ", "lon": 119.30, "lat": 26.08, "type": "OP",  "severity": "轻度"},  # 电镀/有机
    # --- 复合 HMOP × 6 ---
    {"province": "河南",   "city": "济源",   "code": "HA", "lon": 112.60, "lat": 35.07, "type": "HMOP", "severity": "重度"},  # Pb/Cd + 焦化 BaP
    {"province": "陕西",   "city": "宝鸡",   "code": "SN", "lon": 107.24, "lat": 34.36, "type": "HMOP", "severity": "中度"},  # Pb/Zn + TPH
    {"province": "山东",   "city": "济南",   "code": "SD", "lon": 117.00, "lat": 36.65, "type": "HMOP", "severity": "中度"},  # 重金属 + 石化
    {"province": "湖南",   "city": "株洲",   "code": "HN", "lon": 113.13, "lat": 27.83, "type": "HMOP", "severity": "重度"},  # 冶炼 Cd/Pb + 化工
    {"province": "广东",   "city": "广州",   "code": "GD", "lon": 113.26, "lat": 23.13, "type": "HMOP", "severity": "轻度"},  # 工业搬迁地
    {"province": "浙江",   "city": "宁波",   "code": "ZJ", "lon": 121.55, "lat": 29.87, "type": "HMOP", "severity": "轻度"},  # 港口工业
]


# -----------------------------------------------------------------------------
# 采样点配置
# -----------------------------------------------------------------------------
def assign_n_points() -> None:
    """给每个场地分配 80-120 之间稳定的点数（基于索引）"""
    rng = random.Random(SEED + 7)
    for i, site in enumerate(SITE_MATRIX):
        # 每个文件 80-120，避免重复观感
        site["n_points"] = rng.randint(80, 120)


# -----------------------------------------------------------------------------
# 采样点编号 / 区域生成
# -----------------------------------------------------------------------------
REGION_NAMES: Dict[str, List[str]] = {
    "云南": ["西北区", "东南区", "老城区", "尾矿坝区", "冶炼厂区"],
    "湖南": ["中心区", "北部区", "河东区", "冶炼区", "工业带"],
    "广东": ["北区", "南区", "矿区带", "搬迁区", "下游区"],
    "广西": ["龙江区", "矿区带", "城南", "城北", "尾矿区"],
    "江苏": ["栖霞区", "化工园区", "长江沿岸", "城北", "老工业带"],
    "浙江": ["富春江沿岸", "城东区", "工业区", "南区", "港口区"],
    "辽宁": ["铁西区", "化工区", "皇姑区", "新城子", "抚顺接壤带"],
    "山东": ["张店区", "齐鲁化工园", "高新区", "周村区", "临淄区"],
    "湖北": ["青山区", "化工新区", "东湖高新", "阳逻", "葛店"],
    "四川": ["青白江区", "新都区", "化工路", "龙泉驿", "金堂"],
    "福建": ["马尾区", "仓山区", "晋安区", "长乐", "福清"],
    "河南": ["济源城区", "玉川产业区", "梨林镇", "五龙口", "轵城"],
    "陕西": ["金台区", "陈仓区", "高新开发区", "凤翔", "岐山"],
}


def gen_sample_id(code: str, idx: int) -> str:
    """生成采样点编号，例如 YN-001"""
    return f"{code}-{idx:03d}"


# -----------------------------------------------------------------------------
# 数值生成核心
# -----------------------------------------------------------------------------
def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def gen_pH(rng: random.Random) -> float:
    """大部分 6.0-8.0，少量 4.5-5.5"""
    if rng.random() < 0.12:  # ~12% 偏酸
        return round(rng.uniform(4.5, 5.5), 2)
    return round(rng.uniform(6.0, 8.0), 2)


def gen_nutrient(rng: random.Random) -> Dict[str, float]:
    """土壤养分因子，正常范围"""
    return {
        "有机质(g/kg)": round(rng.gauss(22.0, 7.0), 2),       # 10-40
        "全氮(g/kg)": round(rng.gauss(1.5, 0.4), 3),          # 1-3
        "全磷(g/kg)": round(rng.gauss(0.9, 0.3), 3),          # 0.5-1.5
        "全钾(g/kg)": round(rng.gauss(19.0, 3.5), 2),         # 15-25
        "碱解氮(mg/kg)": round(rng.gauss(85.0, 30.0), 1),     # 30-150
        "速效磷(mg/kg)": round(rng.gauss(18.0, 9.0), 2),      # 5-40
        "速效钾(mg/kg)": round(rng.gauss(130.0, 45.0), 1),    # 50-250
    }


def gen_hm_value(rng: random.Random, factor: str, exceed: bool, mult_range: Tuple[float, float]) -> float:
    """生成重金属因子值"""
    thr = HM_THRESHOLD[factor]
    bg_mean, bg_std = HM_BACKGROUND[factor]
    if not exceed:
        # 背景附近
        v = rng.gauss(bg_mean, bg_std)
        v = clamp(v, bg_mean * 0.3, thr * 0.9)  # 不超过筛选值
        return round(v, 3)
    else:
        m = rng.uniform(*mult_range)
        # 超标倍数 × 筛选值，加入 ±15% 抖动
        v = thr * m * rng.uniform(0.85, 1.15)
        return round(v, 3)


def gen_op_value(rng: random.Random, factor: str, exceed: bool, mult_range: Tuple[float, float]) -> float:
    """生成有机物因子值"""
    thr = OP_THRESHOLD[factor]
    bg_mean, bg_std = OP_BACKGROUND[factor]
    if not exceed:
        # 背景/痕量
        if rng.random() < 0.4:  # 40% 未检出
            return round(bg_mean * 0.3, 3)
        v = rng.gauss(bg_mean, bg_std)
        v = clamp(v, 0.0, thr * 0.9)
        return round(max(v, 0.001), 3)
    else:
        m = rng.uniform(*mult_range)
        v = thr * m * rng.uniform(0.85, 1.15)
        return round(v, 3)


def pick_exceed_factors(
    rng: random.Random,
    poll_type: str,
    primary_factors: List[str],
    exceed_ratio: float,
) -> List[str]:
    """
    决定本次场地主超标因子集合（固定 2-4 个主因子，每个点位的超标在这些因子中抽样）。
    返回每个点位"会触发超标"的因子清单。
    """
    # 主因子数：轻度取 2，中度取 3，重度取 4
    n_main = {"轻度": 2, "中度": 3, "重度": 4}
    k = min(n_main.get(exceed_ratio_name(exceed_ratio), 3), len(primary_factors))
    # 实际上 exceed_ratio 直接映射严重度名称
    return rng.sample(primary_factors, k)


def exceed_ratio_name(ratio: float) -> str:
    """根据超标比例反查严重度名（用于确定主因子个数）"""
    for name, cfg in SEVERITY_CFG.items():
        if abs(cfg["ratio"] - ratio) < 1e-6:
            return name
    return "中度"


# -----------------------------------------------------------------------------
# 构建一份场地的全部采样数据
# -----------------------------------------------------------------------------
def build_site_rows(site: Dict) -> Tuple[List[List], Dict]:
    """
    返回 (rows, meta)。
    rows：每行对应一个采样点，元素顺序与 COLUMNS 一致。
    meta：验证信息（采样点数、主超标因子、超标倍数范围、经纬度范围）。
    """
    rng = random.Random(SEED + hash_str(f"{site['province']}{site['city']}{site['type']}{site['severity']}"))
    province = site["province"]
    code = site["code"]
    poll_type = site["type"]
    sev = site["severity"]
    n = site["n_points"]

    cfg = SEVERITY_CFG[sev]
    exceed_ratio = cfg["ratio"]
    mult_range = cfg["mult"]

    # 主超标因子候选
    if poll_type == "HM":
        primary_pool = list(HM_THRESHOLD.keys())
    elif poll_type == "OP":
        primary_pool = list(OP_THRESHOLD.keys())
    else:  # HMOP
        primary_pool = list(HM_THRESHOLD.keys()) + list(OP_THRESHOLD.keys())

    main_factors = pick_exceed_factors(rng, poll_type, primary_pool, exceed_ratio)

    # 区域
    regions = REGION_NAMES.get(province, ["区域A", "区域B", "区域C", "区域D"])

    # 经纬度抖动
    lon0, lat0 = site["lon"], site["lat"]

    rows: List[List] = []
    n_exceed = 0  # 至少一个主因子超标的点位数

    for i in range(1, n + 1):
        # 决定该点位是否"超标点"（按比例）
        is_exceed_point = rng.random() < exceed_ratio
        if is_exceed_point:
            n_exceed += 1

        # 经纬度抖动
        lon = round(lon0 + rng.uniform(-0.30, 0.30), 6)
        lat = round(lat0 + rng.uniform(-0.20, 0.20), 6)

        # 深度（多为表层 0-20cm，少量分层）
        depth_choices = [(0, 20), (0, 50), (20, 50), (50, 100)]
        d_lo, d_hi = rng.choices(depth_choices, weights=[0.55, 0.2, 0.15, 0.1])[0]

        # 土壤类型
        soil_type = {
            "HM": "重金属污染场地",
            "OP": "有机污染场地",
            "HMOP": "复合污染场地",
        }[poll_type]

        pH = gen_pH(rng)
        nutrient = gen_nutrient(rng)

        # 重金属
        hm_vals: Dict[str, float] = {}
        for f in HM_THRESHOLD:
            # 该因子是否参与超标：复合/HM 场地且属于 main_factors
            should_exceed = is_exceed_point and (f in main_factors)
            hm_vals[f] = gen_hm_value(rng, f, should_exceed, mult_range)

        # 有机物
        op_vals: Dict[str, float] = {}
        for f in OP_THRESHOLD:
            should_exceed = is_exceed_point and (f in main_factors)
            op_vals[f] = gen_op_value(rng, f, should_exceed, mult_range)

        # 区域（从池中随机选，保证多样）
        region = rng.choice(regions)

        # 备注
        if is_exceed_point:
            exceed_list = [f.split("(")[0] for f, v in {**hm_vals, **op_vals}.items()
                           if f in main_factors and v > ({**HM_THRESHOLD, **OP_THRESHOLD}[f])]
            remark = "超标：" + "/".join(exceed_list) if exceed_list else ""
        else:
            remark = ""

        # 组装行（按 COLUMNS 顺序）
        row = [
            i,                                  # 序号
            gen_sample_id(code, i),             # 采样点编号
            lon,                                # 经度
            lat,                                # 纬度
            region,                             # 区域
            d_lo,                               # 深度_上限
            d_hi,                               # 深度_下限
            soil_type,                          # 土壤类型
            pH,                                 # pH
            nutrient["有机质(g/kg)"],
            nutrient["全氮(g/kg)"],
            nutrient["全磷(g/kg)"],
            nutrient["全钾(g/kg)"],
            nutrient["碱解氮(mg/kg)"],
            nutrient["速效磷(mg/kg)"],
            nutrient["速效钾(mg/kg)"],
            hm_vals["镉_Cd(mg/kg)"],
            hm_vals["铅_Pb(mg/kg)"],
            hm_vals["砷_As(mg/kg)"],
            hm_vals["铜_Cu(mg/kg)"],
            hm_vals["锌_Zn(mg/kg)"],
            hm_vals["铬_Cr(mg/kg)"],
            hm_vals["汞_Hg(mg/kg)"],
            hm_vals["镍_Ni(mg/kg)"],
            op_vals["苯并芘_BaP(mg/kg)"],
            op_vals["六六六_HCHs(mg/kg)"],
            op_vals["滴滴涕_DDTs(mg/kg)"],
            op_vals["石油烃_TPH(mg/kg)"],
            op_vals["多氯联苯_PCI(mg/kg)"],
            remark,
        ]
        rows.append(row)

    # 计算每个主因子的实际超标倍数（最大值/筛选值）
    main_exceed_stats: Dict[str, Tuple[float, float]] = {}
    for f in main_factors:
        thr = {**HM_THRESHOLD, **OP_THRESHOLD}[f]
        col_idx = COLUMNS.index(f)
        vals = [r[col_idx] for r in rows]
        max_v = max(vals)
        main_exceed_stats[f] = (round(max_v, 3), round(max_v / thr, 1))

    # 经纬度范围
    lons = [r[2] for r in rows]
    lats = [r[3] for r in rows]

    meta = {
        "n_points": n,
        "n_exceed_points": n_exceed,
        "exceed_ratio_actual": round(n_exceed / n, 2),
        "main_factors": main_factors,
        "main_exceed_stats": main_exceed_stats,
        "lon_range": (round(min(lons), 4), round(max(lons), 4)),
        "lat_range": (round(min(lats), 4), round(max(lats), 4)),
    }

    return rows, meta


def hash_str(s: str) -> int:
    """稳定字符串哈希，用于种子"""
    h = 0
    for ch in s:
        h = (h * 31 + ord(ch)) & 0x7FFFFFFF
    return h


# -----------------------------------------------------------------------------
# 写 Excel
# -----------------------------------------------------------------------------
TYPE_CN = {"HM": "重金属", "OP": "有机", "HMOP": "复合"}
SEV_ORDER = {"轻度": 1, "中度": 2, "重度": 3}


def write_excel(out_path: str, site: Dict, rows: List[List]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    title = f"{site['province']}{site['city']}{TYPE_CN[site['type']]}污染场地（{site['severity']}）完整数据"
    ws.title = title[:31]  # Excel sheet name 限制 31 字符

    # 表头
    ws.append(COLUMNS)

    # 数据
    for row in rows:
        ws.append(row)

    # 样式：表头加粗 + 浅灰底 + 居中
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # 数据行居中（数值）
    for r in range(2, ws.max_row + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=col_idx).alignment = center

    # 列宽
    widths = {
        "序号": 6, "采样点编号": 12, "经度": 12, "纬度": 12, "区域": 12,
        "深度_上限(cm)": 12, "深度_下限(cm)": 12, "土壤类型": 14, "pH": 8,
        "备注": 22,
    }
    for idx, name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(name, 14)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


# -----------------------------------------------------------------------------
# 主流程
# -----------------------------------------------------------------------------
def main() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    assign_n_points()

    type_count = {"HM": 0, "OP": 0, "HMOP": 0}
    sev_count = {"轻度": 0, "中度": 0, "重度": 0}
    province_set = set()

    print("=" * 100)
    print("生成 18 份模拟污染场地数据 Excel 文件")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 100)

    summary_rows: List[Dict] = []

    for site in SITE_MATRIX:
        province_set.add(site["province"])
        type_count[site["type"]] += 1
        sev_count[site["severity"]] += 1

        rows, meta = build_site_rows(site)

        # 文件名
        fname = f"演示_{site['province']}_{TYPE_CN[site['type']]}_{site['severity']}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, fname)
        write_excel(out_path, site, rows)

        summary_rows.append({
            "site": site, "meta": meta, "fname": fname, "out_path": out_path,
        })

    # 打印验证信息
    print(f"\n{'序号':<4}{'文件名':<32}{'采样点':<7}{'超标点':<7}{'超标率':<7}{'主要超标因子':<28}{'超标倍数(最大)':<14}{'经度范围':<22}{'纬度范围':<18}")
    print("-" * 150)
    for i, s in enumerate(summary_rows, 1):
        site = s["site"]
        meta = s["meta"]
        # 主要超标因子（取前 3 个）
        factor_strs = []
        for f in meta["main_factors"][:4]:
            max_v, mult = meta["main_exceed_stats"][f]
            factor_strs.append(f"{f.split('(')[0]}×{mult}")
        factors_str = " ".join(factor_strs)
        lon_str = f"{meta['lon_range'][0]}~{meta['lon_range'][1]}"
        lat_str = f"{meta['lat_range'][0]}~{meta['lat_range'][1]}"
        print(f"{i:<4}{s['fname']:<32}{meta['n_points']:<7}{meta['n_exceed_points']:<7}"
              f"{meta['exceed_ratio_actual']:<7.0%}{factors_str:<30}{'':<14}{lon_str:<22}{lat_str:<18}")

    # 覆盖矩阵统计
    print("\n" + "=" * 100)
    print("覆盖矩阵验证：")
    print(f"  文件总数：{len(SITE_MATRIX)}")
    print(f"  污染类型分布：HM={type_count['HM']}, OP={type_count['OP']}, HMOP={type_count['HMOP']}")
    print(f"  严重度分布：轻度={sev_count['轻度']}, 中度={sev_count['中度']}, 重度={sev_count['重度']}")
    print(f"  省份覆盖（{len(province_set)} 个）：{'、'.join(sorted(province_set))}")
    print("=" * 100)
    print(f"完成！18 份文件已保存到：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()

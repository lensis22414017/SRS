"""Round10 D18-D25 数据契约回归。"""
from __future__ import annotations

import pytest
from pydantic import ValidationError


def _indicator(code: str, value: float = 1.0, **overrides):
    from app.api.economic import EconomicIndicatorInput
    payload = {
        "indicator_code": code, "value": value,
        "source_type": "site_actual", "is_proxy": False,
    }
    payload.update(overrides)
    return EconomicIndicatorInput(**payload)


def test_annual_units_are_converted_to_canonical_units():
    from app.services.economic_units import standardize_unit
    assert standardize_unit(1500, "元/公顷·年", "元/亩·年") == 100
    assert standardize_unit(100, "元/亩", "元/亩·年") == 100
    assert standardize_unit(500, "kg/亩", "kg/公顷·年") == 7500


def test_proxy_source_contract_is_strict():
    with pytest.raises(ValidationError):
        _indicator("D18", source_type="regional_official_proxy", is_proxy=False)
    with pytest.raises(ValidationError):
        _indicator("D18", source_type="test_fixture", is_proxy=True)
    valid = _indicator(
        "D18", source_type="regional_official_proxy", is_proxy=True,
        source_name="某市统计公报", source_url="https://example.org/report",
        source_year=2024, source_geography="某市",
    )
    assert valid.is_proxy is True


def test_duplicate_indicator_codes_are_rejected():
    from app.api.economic import EconomicDataBody
    with pytest.raises(ValidationError):
        EconomicDataBody(evaluation_year=2024, indicators=[_indicator("D18"), _indicator("D18")])


def test_d21_d22_d23_d25_cross_checks():
    from app.api.economic import _validate_cross_check
    raw = {
        "area_hectare": 2.0, "yield_kg": 14000.0,
        "gross_output_yuan": 40000.0, "total_cost_yuan": 32000.0,
        "d21_seed_cost": 600.0, "d21_fertilizer_cost": 1200.0,
        "d21_manure_cost": 300.0, "d21_pesticide_cost": 600.0,
        "d21_film_cost": 300.0,
    }
    _validate_cross_check({"D21": 100.0, "D22": 20000.0, "D23": 1.25, "D25": 7000.0}, raw)
    with pytest.raises(ValueError, match="D21"):
        _validate_cross_check({"D21": 200.0}, raw)
    with pytest.raises(ValueError, match="D22"):
        _validate_cross_check({"D22": 10000.0}, raw)
    with pytest.raises(ValueError, match="D23"):
        _validate_cross_check({"D23": 0.8}, raw)
    with pytest.raises(ValueError, match="D25"):
        _validate_cross_check({"D25": 3000.0}, raw)


def test_partial_raw_inputs_and_nonpositive_denominators_are_rejected():
    from app.api.economic import _validate_cross_check
    with pytest.raises(ValueError, match="面积"):
        _validate_cross_check({"D25": 1.0}, {"yield_kg": 100.0})
    with pytest.raises(ValueError, match="总成本"):
        _validate_cross_check({"D23": 1.0}, {"gross_output_yuan": 100.0, "total_cost_yuan": 0.0})
    with pytest.raises(ValueError, match="D21"):
        _validate_cross_check({"D21": 1.0}, {"area_hectare": 1.0, "d21_seed_cost": 10.0})


def test_authenticated_template_import_get_and_delete_round_trip():
    import io
    import openpyxl
    from fastapi.testclient import TestClient

    from app.db.bootstrap import main as bootstrap
    from app.db.session import SessionLocal
    from app.main import app
    from app.models import Site

    bootstrap()
    db = SessionLocal()
    try:
        site = Site(
            site_code="SRS-ECONAPI",
            name="经济接口闭环场地",
            pollution_type="heavy_metal",
            land_use_type="生产用地",
        )
        db.add(site)
        db.commit()
        db.refresh(site)

        client = TestClient(app)
        login = client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": "Demo@2026"},
        )
        assert login.status_code == 200, login.text
        headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

        template = client.get(
            f"/api/v1/sites/{site.id}/economic-data/template",
            headers=headers,
        )
        assert template.status_code == 200, template.text
        workbook = openpyxl.load_workbook(io.BytesIO(template.content))
        sheet = workbook.active
        header_map = {
            cell.value: index + 1
            for index, cell in enumerate(sheet[1])
        }
        values = {
            "D18": 467.41,
            "D19": 200.54,
            "D20": 244.05,
            "D21": 278.53,
            "D22": 19537.65,
            "D23": 1.039083,
            "D24": 17131.0,
            "D25": 7016.85,
        }
        raw = {
            "面积(公顷)": 1.0,
            "总产量(kg)": 7016.85,
            "总产值(元)": 19537.65,
            "总成本(元)": 18802.8,
            "种子总成本(元)": 1014.6,
            "化肥总成本(元)": 2043.3,
            "农家肥总成本(元)": 145.35,
            "农药总成本(元)": 911.85,
            "农膜总成本(元)": 62.85,
        }
        for row_index in range(2, 10):
            code = sheet.cell(row_index, header_map["指标代码"]).value
            sheet.cell(row_index, header_map["评价年份"], 2020)
            sheet.cell(row_index, header_map["场景"], "production")
            sheet.cell(row_index, header_map["作物/用地"], "水稻")
            sheet.cell(row_index, header_map["数值"], values[code])
            for column, value in raw.items():
                sheet.cell(row_index, header_map[column], value)

        output = io.BytesIO()
        workbook.save(output)
        uploaded = client.post(
            f"/api/v1/sites/{site.id}/economic-data/import",
            headers=headers,
            files={
                "file": (
                    "economic.xlsx",
                    output.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200, uploaded.text
        assert uploaded.json()["economic_complete"] is True
        assert uploaded.json()["indicators_saved"] == 8

        loaded = client.get(
            f"/api/v1/sites/{site.id}/economic-data",
            headers=headers,
        )
        assert loaded.status_code == 200, loaded.text
        body = loaded.json()
        assert len(body["indicators"]) == 8
        assert len(body["raw_inputs"]) == 1
        assert {item["year"] for item in body["indicators"]} == {2020}
        assert {item["scenario"] for item in body["indicators"]} == {"production"}

        deleted = client.delete(
            f"/api/v1/sites/{site.id}/economic-data?year=2020&scenario=production",
            headers=headers,
        )
        assert deleted.status_code == 200, deleted.text
        assert deleted.json()["deleted_indicators"] == 8
        assert deleted.json()["deleted_raw_inputs"] == 1
    finally:
        db.close()
